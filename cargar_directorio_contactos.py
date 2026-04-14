"""
Script para cargar datos del Excel 'Información - Personas por organismo .xlsx'
a la colección 'directorio_contactos' en Firestore.
"""
import openpyxl
from app.firebase_config import db

EXCEL_FILE = "Información - Personas por organismo .xlsx"
COLLECTION = "directorio_contactos"

def cargar_datos():
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb[wb.sheetnames[0]]

    # Primera fila = encabezados
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    print(f"Columnas: {headers}")

    total = 0
    errores = 0
    batch = db.batch()
    batch_count = 0

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        # Saltar filas completamente vacías
        if all(v is None for v in row):
            continue

        doc = {}
        for header, value in zip(headers, row):
            if header is None:
                continue
            # Convertir números de teléfono y cédula a string
            if header in ("telefono", "cedula") and value is not None:
                doc[header] = str(int(value)) if isinstance(value, float) else str(value)
            else:
                doc[header] = value if value is not None else ""

        ref = db.collection(COLLECTION).document()
        batch.set(ref, doc)
        batch_count += 1
        total += 1

        # Firestore batch limit: 500 operaciones
        if batch_count >= 400:
            batch.commit()
            print(f"  Batch commit: {total} registros...")
            batch = db.batch()
            batch_count = 0

    # Commit del último batch
    if batch_count > 0:
        batch.commit()

    print(f"\nCarga completada: {total} registros en '{COLLECTION}'")

if __name__ == "__main__":
    cargar_datos()
